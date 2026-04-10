from __future__ import annotations

from io import BytesIO
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import pytest
from openpyxl import Workbook

from expense_record.models import StatementImportRow
from expense_record.statement_import import (
    UnsupportedStatementFileError,
    detect_statement_source,
    import_statement_rows,
    statement_rows_to_review_rows,
)


def test_detect_statement_source_identifies_wechat_xlsx_bytes():
    source = detect_statement_source("wechat.xlsx", _wechat_fixture_bytes())

    assert source == "wechat"


def test_detect_statement_source_identifies_alipay_csv_bytes():
    source = detect_statement_source("alipay.csv", _alipay_fixture_bytes())

    assert source == "alipay"


def test_detect_statement_source_rejects_malformed_xlsx_bytes():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        detect_statement_source("wechat.xlsx", b"not a zip archive")


def test_detect_statement_source_rejects_malformed_csv_bytes():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        detect_statement_source("alipay.csv", b"\xff\xfe\x00\x00")


def test_detect_statement_source_rejects_wechat_missing_source_marker_with_valid_full_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        detect_statement_source("wechat.xlsx", _wechat_missing_marker_fixture_bytes())


def test_detect_statement_source_rejects_alipay_missing_source_marker_with_valid_full_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        detect_statement_source("alipay.csv", _alipay_missing_marker_fixture_bytes())


def test_detect_statement_source_rejects_wechat_generic_marker_with_valid_full_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        detect_statement_source("wechat.xlsx", _wechat_generic_marker_fixture_bytes())


def test_detect_statement_source_rejects_alipay_generic_marker_with_valid_full_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        detect_statement_source("alipay.csv", _alipay_generic_marker_fixture_bytes())


def test_detect_statement_source_rejects_wechat_marker_with_header_subset_layout():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        detect_statement_source("wechat.xlsx", _wechat_marker_subset_fixture_bytes())


def test_import_statement_rows_normalizes_wechat_rows():
    rows = import_statement_rows("wechat.xlsx", _wechat_fixture_bytes())

    assert rows == [
        StatementImportRow(
            transaction_time="2026-03-29 18:44:00",
            counterparty="叫了个炸鸡",
            direction="支出",
            amount="26.50",
        ),
        StatementImportRow(
            transaction_time="2026-03-29 18:41:00",
            counterparty="商户_沈菊",
            direction="支出",
            amount="10.00",
        ),
    ]


def test_import_statement_rows_normalizes_realistic_wechat_export_shape():
    rows = import_statement_rows("wechat.xlsx", _wechat_real_header_fixture_bytes())

    assert rows == [
        StatementImportRow(
            transaction_time="2026-03-31 08:28:10",
            counterparty="刘记钢材丨通州湾",
            direction="支出",
            amount="7.00",
        ),
        StatementImportRow(
            transaction_time="2026-03-30 08:31:56",
            counterparty="累了，王前花，小吃",
            direction="支出",
            amount="5.00",
        ),
    ]


def test_import_statement_rows_skips_wechat_refund_status_rows():
    rows = import_statement_rows("wechat.xlsx", _wechat_refund_status_fixture_bytes())

    assert rows == [
        StatementImportRow(
            transaction_time="2026-03-29 18:44:00",
            counterparty="叫了个炸鸡",
            direction="支出",
            amount="26.50",
        ),
        StatementImportRow(
            transaction_time="2026-03-29 18:41:00",
            counterparty="商户_沈菊",
            direction="支出",
            amount="10.00",
        ),
    ]


def test_import_statement_rows_normalizes_alipay_rows():
    rows = import_statement_rows("alipay.csv", _alipay_fixture_bytes())

    assert rows == [
        StatementImportRow(
            transaction_time="2026-04-03 18:40:31",
            counterparty="淘宝闪购",
            direction="支出",
            amount="25.40",
        ),
        StatementImportRow(
            transaction_time="2026-04-05 04:08:45",
            counterparty="中欧基金管理有限公司",
            direction="不计收支",
            amount="0.02",
        ),
    ]


def test_import_statement_rows_normalizes_realistic_alipay_export_shape():
    rows = import_statement_rows("alipay.csv", _alipay_real_header_fixture_bytes())

    assert rows == [
        StatementImportRow(
            transaction_time="2026-04-05 04:08:45",
            counterparty="中欧基金管理有限公司",
            direction="不计收支",
            amount="0.02",
        ),
        StatementImportRow(
            transaction_time="2026-04-03 18:40:31",
            counterparty="淘宝闪购",
            direction="支出",
            amount="25.40",
        ),
    ]


def test_statement_import_row_to_dict_returns_api_shape():
    row = StatementImportRow(
        transaction_time="2026-03-29 18:44:00",
        counterparty="叫了个炸鸡",
        direction="支出",
        amount="26.50",
    )

    assert row.to_dict() == {
        "transaction_time": "2026-03-29 18:44:00",
        "counterparty": "叫了个炸鸡",
        "direction": "支出",
        "amount": "26.50",
    }


def test_statement_rows_to_review_rows_adds_blank_required_fields():
    rows = statement_rows_to_review_rows(
        [
            StatementImportRow(
                transaction_time="2026-04-10 10:00:00",
                counterparty="Salary",
                direction="收入",
                amount="5000.00",
            ),
            StatementImportRow(
                transaction_time="2026-04-10 12:00:00",
                counterparty="Lunch",
                direction="支出",
                amount="26.50",
            ),
        ],
        source="wechat",
    )

    assert [row.to_dict() for row in rows] == [
        {
            "date": "2026-04-10 10:00:00",
            "description": "Salary",
            "amount": "+5000.00",
            "direction": "income",
            "category": "",
            "member": "",
            "source": "wechat",
            "entry_type": "income",
            "note": "",
        },
        {
            "date": "2026-04-10 12:00:00",
            "description": "Lunch",
            "amount": "-26.50",
            "direction": "expense",
            "category": "",
            "member": "",
            "source": "wechat",
            "entry_type": "expense",
            "note": "",
        },
    ]


def test_statement_rows_to_review_rows_rejects_invalid_direction():
    with pytest.raises(ValueError, match="Invalid direction."):
        statement_rows_to_review_rows(
            [
                StatementImportRow(
                    transaction_time="2026-04-10 10:00:00",
                    counterparty="Mystery",
                    direction="退款",
                    amount="10.00",
                )
            ],
            source="wechat",
        )


def test_import_statement_rows_skips_wechat_footer_row_after_detail_table():
    rows = import_statement_rows("wechat.xlsx", _wechat_with_footer_fixture_bytes())

    assert rows == [
        StatementImportRow(
            transaction_time="2026-03-29 18:44:00",
            counterparty="叫了个炸鸡",
            direction="支出",
            amount="26.50",
        ),
        StatementImportRow(
            transaction_time="2026-03-29 18:41:00",
            counterparty="商户_沈菊",
            direction="支出",
            amount="10.00",
        ),
    ]


def test_import_statement_rows_skips_alipay_footer_row_after_detail_table():
    rows = import_statement_rows("alipay.csv", _alipay_with_footer_fixture_bytes())

    assert rows == [
        StatementImportRow(
            transaction_time="2026-04-03 18:40:31",
            counterparty="淘宝闪购",
            direction="支出",
            amount="25.40",
        ),
        StatementImportRow(
            transaction_time="2026-04-05 04:08:45",
            counterparty="中欧基金管理有限公司",
            direction="不计收支",
            amount="0.02",
        ),
    ]


def test_import_statement_rows_rejects_malformed_wechat_transaction_row_after_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("wechat.xlsx", _wechat_malformed_transaction_fixture_bytes())


def test_import_statement_rows_rejects_malformed_alipay_transaction_row_after_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("alipay.csv", _alipay_malformed_transaction_fixture_bytes())


def test_import_statement_rows_rejects_shortened_wechat_detail_row_after_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("wechat.xlsx", _wechat_shortened_transaction_fixture_bytes())


def test_import_statement_rows_rejects_shortened_wechat_text_timestamp_row_after_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("wechat.xlsx", _wechat_shortened_text_timestamp_fixture_bytes())


def test_import_statement_rows_rejects_shortened_alipay_detail_row_after_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("alipay.csv", _alipay_shortened_transaction_fixture_bytes())


def test_import_statement_rows_rejects_unsupported_file():
    with pytest.raises(UnsupportedStatementFileError):
        import_statement_rows("notes.txt", b"not a statement")


def test_import_statement_rows_rejects_malformed_xlsx_bytes():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("wechat.xlsx", b"not a zip archive")


def test_import_statement_rows_rejects_malformed_csv_bytes():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("alipay.csv", b"\xff\xfe\x00\x00")


def test_import_statement_rows_rejects_wechat_missing_source_marker_with_valid_full_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("wechat.xlsx", _wechat_missing_marker_fixture_bytes())


def test_import_statement_rows_rejects_alipay_missing_source_marker_with_valid_full_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("alipay.csv", _alipay_missing_marker_fixture_bytes())


def test_import_statement_rows_rejects_wechat_generic_marker_with_valid_full_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("wechat.xlsx", _wechat_generic_marker_fixture_bytes())


def test_import_statement_rows_rejects_alipay_generic_marker_with_valid_full_header():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("alipay.csv", _alipay_generic_marker_fixture_bytes())


def test_import_statement_rows_rejects_invalid_alipay_timestamp_with_colons():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("alipay.csv", _alipay_invalid_timestamp_fixture_bytes())


def test_import_statement_rows_rejects_corrupted_xlsx_shared_string_reference():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("wechat.xlsx", _corrupted_shared_string_reference_fixture_bytes())


def test_import_statement_rows_rejects_wechat_marker_with_header_subset_layout():
    with pytest.raises(UnsupportedStatementFileError, match="Unsupported or ambiguous statement file."):
        import_statement_rows("wechat.xlsx", _wechat_marker_subset_fixture_bytes())


def _wechat_fixture_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "微信账单"
    worksheet.append(["微信支付账单明细", "2026-04-06"])
    worksheet.append(["账单说明", "微信支付账单明细"])
    worksheet.append([])
    worksheet.append(["交易时间", "交易类型", "交易对方", "商品说明", "收/支", "金额(元)"])
    worksheet.append([46110.78055555555, "支付", "叫了个炸鸡", "晚餐", "支出", 26.5])
    worksheet.append([46110.77847222222, "支付", "商户_沈菊", "早餐", "支出", 10])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _wechat_real_header_fixture_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "微信账单"
    worksheet.append(["微信支付账单明细"])
    worksheet.append(["导出时间", "2026-04-06"])
    worksheet.append(["----------------------微信支付账单明细列表--------------------"])
    worksheet.append(
        [
            "交易时间",
            "交易类型",
            "交易对方",
            "商品",
            "收/支",
            "金额(元)",
            "支付方式",
            "当前状态",
            "交易单号",
            "商户单号",
            "备注",
        ]
    )
    worksheet.append(
        [
            46112.35289351852,
            "扫二维码付款",
            "刘记钢材丨通州湾",
            "收款方备注:二维码收款",
            "支出",
            "7",
            "中国银行储蓄卡(0453)",
            "已转账",
            "53110001469164202603313597977923",
            "10001073012026033100962578298067",
            "/",
        ]
    )
    worksheet.append(
        [
            46111.35550925926,
            "扫二维码付款",
            "累了，王前花，小吃",
            "收款方备注:二维码收款",
            "支出",
            "5",
            "中国银行储蓄卡(7188)",
            "已转账",
            "53110001470032202603300592217773",
            "10001073012026033000362269775191",
            "/",
        ]
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _wechat_refund_status_fixture_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "微信账单"
    worksheet.append(["微信支付账单明细"])
    worksheet.append(["导出时间", "2026-04-06"])
    worksheet.append(["----------------------微信支付账单明细列表--------------------"])
    worksheet.append(
        [
            "交易时间",
            "交易类型",
            "交易对方",
            "商品",
            "收/支",
            "金额(元)",
            "支付方式",
            "当前状态",
            "交易单号",
            "商户单号",
            "备注",
        ]
    )
    worksheet.append(
        [
            46110.78055555555,
            "商户消费",
            "叫了个炸鸡",
            "美团收银909700210917833777",
            "支出",
            "26.5",
            "中信银行信用卡(1709)",
            "支付成功",
            "4200003040202603292690038965",
            "0461368606473567868807150",
            "/",
        ]
    )
    worksheet.append(
        [
            46096.44118055556,
            "商户消费",
            "抖音电商商家",
            "/",
            "支出",
            "401",
            "零钱",
            "已退款(¥398.00)",
            "4200000000000000000000000000",
            "mock-merchant-order",
            "/",
        ]
    )
    worksheet.append(
        [
            46110.77847222222,
            "商户消费",
            "商户_沈菊",
            "付款码支付",
            "支出",
            "10",
            "中信银行信用卡(1709)",
            "支付成功",
            "4200003071202603299937870599",
            "20260329184103346000061W",
            "/",
        ]
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _wechat_with_footer_fixture_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "微信账单"
    worksheet.append(["微信支付账单明细", "2026-04-06"])
    worksheet.append(["账单说明", "微信支付账单明细"])
    worksheet.append([])
    worksheet.append(["交易时间", "交易类型", "交易对方", "商品说明", "收/支", "金额(元)"])
    worksheet.append([46110.78055555555, "支付", "叫了个炸鸡", "晚餐", "支出", 26.5])
    worksheet.append([46110.77847222222, "支付", "商户_沈菊", "早餐", "支出", 10])
    worksheet.append(["共2笔"])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _wechat_malformed_transaction_fixture_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "微信账单"
    worksheet.append(["微信支付账单明细", "2026-04-06"])
    worksheet.append(["账单说明", "微信支付账单明细"])
    worksheet.append([])
    worksheet.append(["交易时间", "交易类型", "交易对方", "商品说明", "收/支", "金额(元)"])
    worksheet.append([46110.78055555555, "支付", "叫了个炸鸡", "晚餐", "支出", "not-an-amount"])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _wechat_shortened_transaction_fixture_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "微信账单"
    worksheet.append(["微信支付账单明细", "2026-04-06"])
    worksheet.append(["账单说明", "微信支付账单明细"])
    worksheet.append([])
    worksheet.append(["交易时间", "交易类型", "交易对方", "商品说明", "收/支", "金额(元)"])
    worksheet.append([46110.78055555555, "支付", "叫了个炸鸡", "晚餐", "支出"])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _wechat_shortened_text_timestamp_fixture_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "微信账单"
    worksheet.append(["微信支付账单明细", "2026-04-06"])
    worksheet.append(["账单说明", "微信支付账单明细"])
    worksheet.append([])
    worksheet.append(["交易时间", "交易类型", "交易对方", "商品说明", "收/支", "金额(元)"])
    worksheet.append(["2026-03-29 18:44:00", "支付", "叫了个炸鸡", "晚餐", "支出"])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _wechat_missing_marker_fixture_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "微信账单"
    worksheet.append(["导出时间", "2026-04-06"])
    worksheet.append(["账单说明", "普通导出"])
    worksheet.append([])
    worksheet.append(["交易时间", "交易类型", "交易对方", "商品说明", "收/支", "金额(元)"])
    worksheet.append([46110.78055555555, "支付", "叫了个炸鸡", "晚餐", "支出", 26.5])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _wechat_generic_marker_fixture_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "微信账单"
    worksheet.append(["微信账单", "2026-04-06"])
    worksheet.append(["账单说明", "微信导出"])
    worksheet.append([])
    worksheet.append(["交易时间", "交易类型", "交易对方", "商品说明", "收/支", "金额(元)"])
    worksheet.append([46110.78055555555, "支付", "叫了个炸鸡", "晚餐", "支出", 26.5])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _wechat_marker_subset_fixture_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "微信账单"
    worksheet.append(["微信支付账单明细", "2026-04-06"])
    worksheet.append(["账单说明", "微信支付账单明细"])
    worksheet.append([])
    worksheet.append(["交易时间", "交易对方", "收/支", "金额(元)"])
    worksheet.append([46110.78055555555, "叫了个炸鸡", "支出", 26.5])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _alipay_fixture_bytes() -> bytes:
    return (
        "支付宝支付科技有限公司\n"
        "账单说明,支付宝账户\n"
        "交易时间,交易分类,交易对方,对方账号,商品说明,收/支,金额,收/付款方式,交易状态,交易订单号,商家订单号,备注\n"
        "2026-04-03 18:40:31,消费,淘宝闪购,支付宝账户,外卖,支出,25.4,余额宝,成功,202604030001,202604030001A,外卖\n"
        "2026-04-05 04:08:45,理财,中欧基金管理有限公司,支付宝账户,基金,不计收支,0.02,余额宝,成功,202604050001,202604050001A,基金\n"
    ).encode("gb18030")


def _alipay_with_footer_fixture_bytes() -> bytes:
    return (
        "支付宝支付科技有限公司\n"
        "账单说明,支付宝账户\n"
        "交易时间,交易分类,交易对方,对方账号,商品说明,收/支,金额,收/付款方式,交易状态,交易订单号,商家订单号,备注\n"
        "2026-04-03 18:40:31,消费,淘宝闪购,支付宝账户,外卖,支出,25.4,余额宝,成功,202604030001,202604030001A,外卖\n"
        "2026-04-05 04:08:45,理财,中欧基金管理有限公司,支付宝账户,基金,不计收支,0.02,余额宝,成功,202604050001,202604050001A,基金\n"
        "共2笔\n"
    ).encode("gb18030")


def _alipay_malformed_transaction_fixture_bytes() -> bytes:
    return (
        "支付宝支付科技有限公司\n"
        "账单说明,支付宝账户\n"
        "交易时间,交易分类,交易对方,对方账号,商品说明,收/支,金额,收/付款方式,交易状态,交易订单号,商家订单号,备注\n"
        "2026-04-03 18:40:31,消费,淘宝闪购,支付宝账户,外卖,支出,not-an-amount,余额宝,成功,202604030001,202604030001A,外卖\n"
    ).encode("gb18030")


def _alipay_shortened_transaction_fixture_bytes() -> bytes:
    return (
        "支付宝支付科技有限公司\n"
        "账单说明,支付宝账户\n"
        "交易时间,交易分类,交易对方,对方账号,商品说明,收/支,金额,收/付款方式,交易状态,交易订单号,商家订单号,备注\n"
        "2026-04-03 18:40:31,消费,淘宝闪购,支付宝账户,外卖,支出\n"
    ).encode("gb18030")


def _alipay_real_header_fixture_bytes() -> bytes:
    return (
        "------------------------支付宝支付科技有限公司  电子客户回单------------------------\n"
        "账单说明,支付宝账户\n"
        "交易时间,交易分类,交易对方,对方账号,商品说明,收/支,金额,收/付款方式,交易状态,交易订单号,商家订单号,备注,\n"
        "2026-04-05 04:08:45,投资理财,中欧基金管理有限公司,/,余额宝-2026.04.04-收益发放,不计收支,0.02,余额宝,交易成功,20260405308428195141\\t,\\t,,\n"
        "2026-04-03 18:40:31,餐饮美食,淘宝闪购,e50***@alibaba-inc.com,隆江猪脚饭(鑫凯隆购物中心店)外卖订单,支出,25.40,招联消金信用购,交易成功,2026040323001113141424091931\\t,13180600726040396248521892229\\t,,\n"
    ).encode("gb18030")


def _alipay_missing_marker_fixture_bytes() -> bytes:
    return (
        "账单说明,电子账单\n"
        "收支明细导出,2026-04-06\n"
        "交易时间,交易分类,交易对方,对方账号,商品说明,收/支,金额,收/付款方式,交易状态,交易订单号,商家订单号,备注\n"
        "2026-04-03 18:40:31,消费,淘宝闪购,支付宝账户,外卖,支出,25.4,余额宝,成功,202604030001,202604030001A,外卖\n"
    ).encode("gb18030")


def _alipay_generic_marker_fixture_bytes() -> bytes:
    return (
        "支付宝账单\n"
        "账单说明,支付宝账户\n"
        "交易时间,交易分类,交易对方,对方账号,商品说明,收/支,金额,收/付款方式,交易状态,交易订单号,商家订单号,备注\n"
        "2026-04-03 18:40:31,消费,淘宝闪购,支付宝账户,外卖,支出,25.4,余额宝,成功,202604030001,202604030001A,外卖\n"
    ).encode("gb18030")


def _alipay_invalid_timestamp_fixture_bytes() -> bytes:
    return (
        "支付宝支付科技有限公司\n"
        "账单说明,支付宝账户\n"
        "交易时间,交易分类,交易对方,对方账号,商品说明,收/支,金额,收/付款方式,交易状态,交易订单号,商家订单号,备注\n"
        "2026-04-03 18:40,消费,淘宝闪购,支付宝账户,外卖,支出,25.4,余额宝,成功,202604030001,202604030001A,外卖\n"
    ).encode("gb18030")


def _corrupted_shared_string_reference_fixture_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "微信账单"
    worksheet.append(["微信支付账单明细", "2026-04-06"])
    worksheet.append(["账单说明", "微信支付账单明细"])
    worksheet.append([])
    worksheet.append(["交易时间", "交易类型", "交易对方", "商品说明", "收/支", "金额(元)"])
    worksheet.append([46110.78055555555, "支付", "叫了个炸鸡", "晚餐", "支出", 26.5])

    buffer = BytesIO()
    workbook.save(buffer)
    original = buffer.getvalue()

    input_zip = BytesIO(original)
    output_zip = BytesIO()
    with ZipFile(input_zip, "r") as zin, ZipFile(output_zip, "w") as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                data = _corrupt_sheet1_shared_string_reference(data)
            zout.writestr(info, data)
    return output_zip.getvalue()


def _corrupt_sheet1_shared_string_reference(sheet_xml: bytes) -> bytes:
    root = ET.fromstring(sheet_xml)
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    cell = root.find(".//main:c[@r='C5']", ns)
    if cell is None:
        raise AssertionError("expected worksheet cell not found")
    for child in list(cell):
        cell.remove(child)
    cell.attrib["t"] = "s"
    value = ET.SubElement(cell, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
    value.text = "999"
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)
