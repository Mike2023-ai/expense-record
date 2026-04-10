from expense_record.models import (
    AssetSnapshot,
    CategoryRecord,
    ExpenseRow,
    LedgerEntry,
    MemberRecord,
    StockRecord,
)
from expense_record.config import DEFAULT_CATEGORIES
from expense_record.storage import ExcelExpenseStorage
from datetime import date, datetime
from openpyxl import Workbook, load_workbook


def test_excel_storage_creates_workbook_and_appends_rows(tmp_path):
    workbook_path = tmp_path / "expenses.xlsx"
    storage = ExcelExpenseStorage(workbook_path)

    storage.append_row(ExpenseRow(date="2026-03-29", merchant_item="星巴克咖啡", amount="32.00"))
    storage.append_row(ExpenseRow(date="2026-03-30", merchant_item="便利店", amount="8.50"))

    assert workbook_path.exists()
    assert storage.list_rows() == [
        ExpenseRow(date="2026-03-29", merchant_item="星巴克咖啡", amount="32.00"),
        ExpenseRow(date="2026-03-30", merchant_item="便利店", amount="8.50"),
    ]


def test_storage_append_rows_appends_multiple_rows(tmp_path):
    workbook_path = tmp_path / "expenses.xlsx"
    storage = ExcelExpenseStorage(workbook_path)

    storage.append_rows(
        [
            ExpenseRow(date="03-28", merchant_item="滴滴出行", amount="28.00"),
            ExpenseRow(date="03-29", merchant_item="早餐", amount="5.00"),
        ]
    )

    assert storage.list_rows() == [
        ExpenseRow(date="03-28", merchant_item="滴滴出行", amount="28.00"),
        ExpenseRow(date="03-29", merchant_item="早餐", amount="5.00"),
    ]


def test_excel_storage_preserves_preexisting_rows_with_custom_header(tmp_path):
    workbook_path = tmp_path / "expenses.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "expenses"
    worksheet.append(["Date", "Merchant", "Amount"])
    worksheet.append(["2026-03-01", "Existing Shop", "11.00"])
    workbook.save(workbook_path)

    storage = ExcelExpenseStorage(workbook_path)

    assert storage.list_rows() == [ExpenseRow(date="2026-03-01", merchant_item="Existing Shop", amount="11.00")]

    storage.append_row(ExpenseRow(date="2026-03-02", merchant_item="New Shop", amount="12.00"))

    reloaded = load_workbook(workbook_path).active
    assert [row for row in reloaded.iter_rows(values_only=True)] == [
        ("Date", "Merchant", "Amount", None),
        ("2026-03-01", "Existing Shop", "11.00", None),
        ("2026-03-02", "New Shop", "12.00", None),
    ]


def test_excel_storage_uses_named_expenses_sheet_in_multi_sheet_workbook(tmp_path):
    workbook_path = tmp_path / "expenses.xlsx"
    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "summary"
    default_sheet.append(["ignore", "these", "rows"])

    expenses_sheet = workbook.create_sheet("expenses")
    expenses_sheet.append(["date", "merchant/item", "amount"])
    expenses_sheet.append(["2026-03-01", "Existing Shop", "11.00"])
    workbook.active = 0
    workbook.save(workbook_path)

    storage = ExcelExpenseStorage(workbook_path)

    assert storage.list_rows() == [ExpenseRow(date="2026-03-01", merchant_item="Existing Shop", amount="11.00")]

    storage.append_row(ExpenseRow(date="2026-03-02", merchant_item="New Shop", amount="12.00"))

    reloaded = load_workbook(workbook_path)
    assert [row for row in reloaded["summary"].iter_rows(values_only=True)] == [("ignore", "these", "rows")]
    assert [row for row in reloaded["expenses"].iter_rows(values_only=True)] == [
        ("date", "merchant/item", "amount", None),
        ("2026-03-01", "Existing Shop", "11.00", None),
        ("2026-03-02", "New Shop", "12.00", None),
    ]


def test_excel_storage_creates_missing_expenses_sheet_with_headers(tmp_path):
    workbook_path = tmp_path / "expenses.xlsx"
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["ignore", "these", "rows"])
    workbook.save(workbook_path)

    storage = ExcelExpenseStorage(workbook_path)

    assert storage.list_rows() == []

    storage.append_row(ExpenseRow(date="2026-03-02", merchant_item="New Shop", amount="12.00"))

    reloaded = load_workbook(workbook_path)
    assert [row for row in reloaded["summary"].iter_rows(values_only=True)] == [("ignore", "these", "rows")]
    assert [row for row in reloaded["expenses"].iter_rows(values_only=True)] == [
        ("date", "merchant/item", "amount", "direction"),
        ("2026-03-02", "New Shop", "12.00", None),
    ]


def test_excel_storage_initializes_blank_existing_expenses_sheet(tmp_path):
    workbook_path = tmp_path / "expenses.xlsx"
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["ignore", "these", "rows"])
    workbook.create_sheet("expenses")
    workbook.save(workbook_path)

    storage = ExcelExpenseStorage(workbook_path)

    assert storage.list_rows() == []

    storage.append_row(ExpenseRow(date="2026-03-03", merchant_item="Blank Sheet Shop", amount="13.00"))

    reloaded = load_workbook(workbook_path)
    assert [row for row in reloaded["summary"].iter_rows(values_only=True)] == [("ignore", "these", "rows")]
    assert [row for row in reloaded["expenses"].iter_rows(values_only=True)] == [
        ("date", "merchant/item", "amount", "direction"),
        ("2026-03-03", "Blank Sheet Shop", "13.00", None),
    ]


def test_excel_storage_normalizes_typed_date_cells(tmp_path):
    workbook_path = tmp_path / "expenses.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "expenses"
    worksheet.append(["date", "merchant/item", "amount"])
    worksheet.append([date(2026, 3, 4), "Typed Date Shop", "14.00"])
    worksheet.append([datetime(2026, 3, 5, 15, 30), "Typed Datetime Shop", "15.00"])
    workbook.save(workbook_path)

    storage = ExcelExpenseStorage(workbook_path)

    assert storage.list_rows() == [
        ExpenseRow(date="2026-03-04", merchant_item="Typed Date Shop", amount="14.00"),
        ExpenseRow(date="2026-03-05", merchant_item="Typed Datetime Shop", amount="15.00"),
    ]


def test_ledger_entry_to_dict_includes_member_category_and_signed_amount():
    entry = LedgerEntry(
        date="2026-04-10",
        description="Salary",
        amount="+5000.00",
        direction="income",
        category="salary",
        member="Mike",
        source="manual",
        entry_type="income",
        note="April salary",
    )

    assert entry.to_dict() == {
        "date": "2026-04-10",
        "description": "Salary",
        "amount": "+5000.00",
        "direction": "income",
        "category": "salary",
        "member": "Mike",
        "source": "manual",
        "entry_type": "income",
        "note": "April salary",
    }


def test_asset_snapshot_to_dict_returns_expected_shape():
    snapshot = AssetSnapshot(
        date="2026-04-30",
        cash_or_balance_total="20000.00",
        stock_total_value="150000.00",
        note="Month end",
    )

    assert snapshot.to_dict() == {
        "date": "2026-04-30",
        "cash_or_balance_total": "20000.00",
        "stock_total_value": "150000.00",
        "note": "Month end",
    }


def test_ledger_entry_docstring_clarifies_direction_and_entry_type():
    assert LedgerEntry.__doc__ == (
        "LedgerEntry represents a family ledger row.\n\n"
        "direction stores the cash flow direction, while entry_type stores the\n"
        "business classification of the row."
    )


def test_category_record_to_dict_returns_expected_shape():
    record = CategoryRecord(
        date="2026-04-30",
        category="salary",
        amount="5000.00",
        direction="income",
        note="April salary",
    )

    assert record.to_dict() == {
        "date": "2026-04-30",
        "category": "salary",
        "amount": "5000.00",
        "direction": "income",
        "note": "April salary",
    }


def test_member_record_to_dict_returns_expected_shape():
    record = MemberRecord(
        date="2026-04-30",
        member="Mike",
        amount="5000.00",
        direction="income",
        note="April salary",
    )

    assert record.to_dict() == {
        "date": "2026-04-30",
        "member": "Mike",
        "amount": "5000.00",
        "direction": "income",
        "note": "April salary",
    }


def test_stock_record_to_dict_returns_expected_shape():
    record = StockRecord(
        date="2026-04-30",
        stock_name="ACME",
        stock_quantity="10",
        stock_price="100.00",
        stock_total_value="1000.00",
        note="Month end",
    )

    assert record.to_dict() == {
        "date": "2026-04-30",
        "stock_name": "ACME",
        "stock_quantity": "10",
        "stock_price": "100.00",
        "stock_total_value": "1000.00",
        "note": "Month end",
    }


def test_storage_persists_manual_ledger_entry_with_member_category_and_signed_amount(tmp_path):
    storage = ExcelExpenseStorage(tmp_path / "family.xlsx")

    storage.append_ledger_entries(
        [
            LedgerEntry(
                date="2026-04-10",
                description="Salary",
                amount="+5000.00",
                direction="income",
                category="salary",
                member="Mike",
                source="manual",
                entry_type="income",
                note="April",
            )
        ]
    )

    assert storage.list_ledger_entries() == [
        LedgerEntry(
            date="2026-04-10",
            description="Salary",
            amount="+5000.00",
            direction="income",
            category="salary",
            member="Mike",
            source="manual",
            entry_type="income",
            note="April",
        )
    ]


def test_storage_replaces_categories_and_members_lists(tmp_path):
    storage = ExcelExpenseStorage(tmp_path / "family.xlsx")

    storage.replace_categories(
        [
            CategoryRecord(
                date="2026-04-10",
                category="salary",
                amount="5000.00",
                direction="income",
                note="Primary income",
            ),
            CategoryRecord(
                date="2026-04-10",
                category="rounding",
                amount="0.99",
                direction="expense",
                note="Should be ignored",
            ),
        ]
    )
    storage.replace_members(
        [
            MemberRecord(
                date="2026-04-10",
                member="Mike",
                amount="5000.00",
                direction="income",
                note="Primary member",
            ),
            MemberRecord(
                date="2026-04-10",
                member="Family",
                amount="-0.50",
                direction="expense",
                note="Should be ignored",
            ),
        ]
    )

    assert storage.list_categories() == [
        CategoryRecord(category="salary"),
        CategoryRecord(category="rounding"),
    ]
    assert storage.list_members() == [
        MemberRecord(member="Mike"),
        MemberRecord(member="Family"),
    ]


def test_storage_list_categories_bootstraps_seeded_defaults_on_first_read(tmp_path):
    storage = ExcelExpenseStorage(tmp_path / "family.xlsx")

    assert [row.category for row in storage.list_categories()] == list(DEFAULT_CATEGORIES)


def test_storage_list_members_bootstraps_empty_sheet_on_first_read(tmp_path):
    storage = ExcelExpenseStorage(tmp_path / "family.xlsx")

    assert storage.list_members() == []


def test_storage_persists_asset_snapshots_and_stock_records(tmp_path):
    storage = ExcelExpenseStorage(tmp_path / "family.xlsx")

    storage.append_asset_snapshots(
        [
            AssetSnapshot(
                date="2026-04-30",
                cash_or_balance_total="20000.00",
                stock_total_value="150000.00",
                note="Month end",
            )
        ]
    )
    storage.append_stock_records(
        [
            StockRecord(
                date="2026-04-30",
                stock_name="ACME",
                stock_quantity="10",
                stock_price="100.00",
                stock_total_value="1000.00",
                note="Brokerage",
            )
        ]
    )

    assert storage.list_asset_snapshots() == [
        AssetSnapshot(
            date="2026-04-30",
            cash_or_balance_total="20000.00",
            stock_total_value="150000.00",
            note="Month end",
        )
    ]
    assert storage.list_stock_records() == [
        StockRecord(
            date="2026-04-30",
            stock_name="ACME",
            stock_quantity="10",
            stock_price="100.00",
            stock_total_value="1000.00",
            note="Brokerage",
        )
    ]


def test_storage_filters_only_ledger_like_amount_rows_below_one(tmp_path):
    storage = ExcelExpenseStorage(tmp_path / "family.xlsx")

    storage.append_ledger_entries(
        [
            LedgerEntry(
                date="2026-04-10",
                description="Interest",
                amount="+0.99",
                direction="income",
                category="interest",
                member="Mike",
                source="manual",
                entry_type="income",
                note="Ignored",
            ),
            LedgerEntry(
                date="2026-04-10",
                description="Bonus",
                amount="+1.00",
                direction="income",
                category="salary",
                member="Mike",
                source="manual",
                entry_type="income",
                note="Kept",
            ),
        ]
    )
    storage.replace_categories(
        [
            CategoryRecord(
                date="2026-04-10",
                category="interest",
                amount="0.50",
                direction="income",
                note="Ignored",
            ),
            CategoryRecord(
                date="2026-04-10",
                category="salary",
                amount="1.00",
                direction="income",
                note="Kept",
            ),
        ]
    )
    storage.replace_members(
        [
            MemberRecord(
                date="2026-04-10",
                member="Child",
                amount="-0.50",
                direction="expense",
                note="Ignored",
            ),
            MemberRecord(
                date="2026-04-10",
                member="Mike",
                amount="-1.00",
                direction="expense",
                note="Kept",
            ),
        ]
    )
    storage.append_asset_snapshots(
        [
            AssetSnapshot(
                date="2026-04-30",
                cash_or_balance_total="0.50",
                stock_total_value="0.00",
                note="Ignored",
            ),
            AssetSnapshot(
                date="2026-04-30",
                cash_or_balance_total="1.00",
                stock_total_value="0.00",
                note="Kept",
            ),
        ]
    )
    storage.append_stock_records(
        [
            StockRecord(
                date="2026-04-30",
                stock_name="PENNY",
                stock_quantity="1",
                stock_price="0.50",
                stock_total_value="0.50",
                note="Ignored",
            ),
            StockRecord(
                date="2026-04-30",
                stock_name="ACME",
                stock_quantity="1",
                stock_price="1.00",
                stock_total_value="1.00",
                note="Kept",
            ),
        ]
    )

    assert storage.list_ledger_entries() == [
        LedgerEntry(
            date="2026-04-10",
            description="Bonus",
            amount="+1.00",
            direction="income",
            category="salary",
            member="Mike",
            source="manual",
            entry_type="income",
            note="Kept",
        )
    ]
    assert storage.list_categories() == [
        CategoryRecord(category="interest"),
        CategoryRecord(category="salary"),
    ]
    assert storage.list_members() == [
        MemberRecord(member="Child"),
        MemberRecord(member="Mike"),
    ]
    assert storage.list_asset_snapshots() == [
        AssetSnapshot(
            date="2026-04-30",
            cash_or_balance_total="0.50",
            stock_total_value="0.00",
            note="Ignored",
        ),
        AssetSnapshot(
            date="2026-04-30",
            cash_or_balance_total="1.00",
            stock_total_value="0.00",
            note="Kept",
        )
    ]
    assert storage.list_stock_records() == [
        StockRecord(
            date="2026-04-30",
            stock_name="PENNY",
            stock_quantity="1",
            stock_price="0.50",
            stock_total_value="0.50",
            note="Ignored",
        ),
        StockRecord(
            date="2026-04-30",
            stock_name="ACME",
            stock_quantity="1",
            stock_price="1.00",
            stock_total_value="1.00",
            note="Kept",
        )
    ]
