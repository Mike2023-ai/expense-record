from __future__ import annotations

from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import date as date_type, datetime as datetime_type
from decimal import Decimal
from pathlib import Path
from typing import Any, TypeVar

from openpyxl import Workbook, load_workbook

from expense_record.models import (
    AssetSnapshot,
    CategoryRecord,
    ExpenseRow,
    LedgerEntry,
    MemberRecord,
    StockRecord,
)

RowT = TypeVar(
    "RowT",
    ExpenseRow,
    LedgerEntry,
    CategoryRecord,
    MemberRecord,
    AssetSnapshot,
    StockRecord,
)


@dataclass(frozen=True, slots=True)
class _SheetSpec:
    name: str
    headers: tuple[str, ...]
    fields: tuple[str, ...]


class ExcelExpenseStorage:
    headers = ("date", "merchant/item", "amount", "direction")
    sheet_name = "expenses"
    ledger_sheet_name = "ledger"
    categories_sheet_name = "categories"
    members_sheet_name = "members"
    asset_snapshots_sheet_name = "asset_snapshots"
    stock_records_sheet_name = "stock_records"

    _EXPENSES_SPEC = _SheetSpec(
        name=sheet_name,
        headers=headers,
        fields=("date", "merchant_item", "amount", "direction"),
    )
    _LEDGER_SPEC = _SheetSpec(
        name=ledger_sheet_name,
        headers=("date", "description", "amount", "direction", "category", "member", "source", "entry_type", "note"),
        fields=("date", "description", "amount", "direction", "category", "member", "source", "entry_type", "note"),
    )
    _CATEGORIES_SPEC = _SheetSpec(
        name=categories_sheet_name,
        headers=("date", "category", "amount", "direction", "note"),
        fields=("date", "category", "amount", "direction", "note"),
    )
    _MEMBERS_SPEC = _SheetSpec(
        name=members_sheet_name,
        headers=("date", "member", "amount", "direction", "note"),
        fields=("date", "member", "amount", "direction", "note"),
    )
    _ASSET_SNAPSHOTS_SPEC = _SheetSpec(
        name=asset_snapshots_sheet_name,
        headers=("date", "cash_or_balance_total", "stock_total_value", "note"),
        fields=("date", "cash_or_balance_total", "stock_total_value", "note"),
    )
    _STOCK_RECORDS_SPEC = _SheetSpec(
        name=stock_records_sheet_name,
        headers=("date", "stock_name", "stock_quantity", "stock_price", "stock_total_value", "note"),
        fields=("date", "stock_name", "stock_quantity", "stock_price", "stock_total_value", "note"),
    )

    def __init__(self, workbook_path: str | Path):
        self.workbook_path = Path(workbook_path)

    def append_row(self, row: ExpenseRow) -> None:
        self.append_rows((row,))

    def append_rows(self, rows: Iterable[ExpenseRow]) -> None:
        self._append_records(self._EXPENSES_SPEC, rows)

    def list_rows(self) -> list[ExpenseRow]:
        return self._list_records(self._EXPENSES_SPEC, ExpenseRow)

    def clear_all(self) -> None:
        self._replace_records(self._EXPENSES_SPEC, ())

    def append_ledger_entries(self, rows: Iterable[LedgerEntry]) -> None:
        self._append_records(self._LEDGER_SPEC, self._filter_ledger_entries(rows))

    def list_ledger_entries(self) -> list[LedgerEntry]:
        return self._list_records(self._LEDGER_SPEC, LedgerEntry)

    def replace_categories(self, rows: Iterable[CategoryRecord]) -> None:
        self._replace_records(self._CATEGORIES_SPEC, self._filter_category_records(rows))

    def list_categories(self) -> list[CategoryRecord]:
        return self._list_records(self._CATEGORIES_SPEC, CategoryRecord)

    def replace_members(self, rows: Iterable[MemberRecord]) -> None:
        self._replace_records(self._MEMBERS_SPEC, self._filter_member_records(rows))

    def list_members(self) -> list[MemberRecord]:
        return self._list_records(self._MEMBERS_SPEC, MemberRecord)

    def append_asset_snapshots(self, rows: Iterable[AssetSnapshot]) -> None:
        self._append_records(self._ASSET_SNAPSHOTS_SPEC, rows)

    def list_asset_snapshots(self) -> list[AssetSnapshot]:
        return self._list_records(self._ASSET_SNAPSHOTS_SPEC, AssetSnapshot)

    def append_stock_records(self, rows: Iterable[StockRecord]) -> None:
        self._append_records(self._STOCK_RECORDS_SPEC, rows)

    def list_stock_records(self) -> list[StockRecord]:
        return self._list_records(self._STOCK_RECORDS_SPEC, StockRecord)

    def _append_records(self, spec: _SheetSpec, rows: Iterable[RowT]) -> None:
        materialized = list(rows)
        if not materialized:
            return

        workbook = self._load_or_create_workbook()
        worksheet = self._get_sheet(workbook, spec)
        for row in materialized:
            worksheet.append([getattr(row, field_name) for field_name in spec.fields])
        workbook.save(self.workbook_path)

    def _replace_records(self, spec: _SheetSpec, rows: Iterable[RowT]) -> None:
        workbook = self._load_or_create_workbook()
        worksheet = self._get_sheet(workbook, spec)
        worksheet.delete_rows(2, worksheet.max_row)
        for row in rows:
            worksheet.append([getattr(row, field_name) for field_name in spec.fields])
        workbook.save(self.workbook_path)

    def _list_records(self, spec: _SheetSpec, model_type: type[RowT]) -> list[RowT]:
        workbook = self._load_or_create_workbook()
        worksheet = self._get_sheet(workbook, spec)
        rows: list[RowT] = []
        expected_columns = len(spec.fields)
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(values):
                continue
            normalized = self._normalize_values(values, expected_columns)
            rows.append(model_type(**dict(zip(spec.fields, normalized, strict=False))))
        return rows

    def _load_or_create_workbook(self):
        if self.workbook_path.exists():
            return load_workbook(self.workbook_path)

        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.sheet_name
        self._write_headers(worksheet, self._EXPENSES_SPEC.headers)
        workbook.save(self.workbook_path)
        return workbook

    def _get_sheet(self, workbook, spec: _SheetSpec):
        if spec.name in workbook.sheetnames:
            worksheet = workbook[spec.name]
            if self._sheet_needs_headers(worksheet):
                self._write_headers(worksheet, spec.headers)
                workbook.save(self.workbook_path)
            return worksheet

        worksheet = workbook.create_sheet(spec.name)
        self._write_headers(worksheet, spec.headers)
        workbook.save(self.workbook_path)
        return worksheet

    def _sheet_needs_headers(self, worksheet) -> bool:
        return worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None

    def _write_headers(self, worksheet, headers: Sequence[str]) -> None:
        for column, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=column, value=header)

    def _normalize_values(self, values: tuple[Any, ...], size: int) -> list[str]:
        normalized: list[str] = []
        for value in list(values)[:size]:
            normalized.append(self._normalize_cell_value(value))
        normalized.extend([""] * (size - len(normalized)))
        return normalized

    def _normalize_cell_value(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime_type):
            return value.date().isoformat()
        if isinstance(value, date_type):
            return value.isoformat()
        return str(value)

    def _filter_ledger_entries(self, rows: Iterable[LedgerEntry]) -> list[LedgerEntry]:
        return [row for row in rows if _is_effective_amount(row.amount)]

    def _filter_category_records(self, rows: Iterable[CategoryRecord]) -> list[CategoryRecord]:
        return [row for row in rows if _is_effective_amount(row.amount)]

    def _filter_member_records(self, rows: Iterable[MemberRecord]) -> list[MemberRecord]:
        return [row for row in rows if _is_effective_amount(row.amount)]

def _is_effective_amount(amount_text: str) -> bool:
    if not amount_text:
        return False
    return abs(Decimal(amount_text)) >= Decimal("1")
